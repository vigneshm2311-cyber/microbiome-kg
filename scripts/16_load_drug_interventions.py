"""
16_load_drug_interventions.py

Populates intervention and microbe_intervention -- both empty since
the schema was designed -- using real data from Maier et al. 2018
(Nature 555:623-628), "Extensive impact of non-antibiotic drugs on
human gut bacteria": 1,197 compounds screened against 40 representative
gut bacterial strains, FDR-adjusted p-values for growth inhibition.

Files used (place in data/raw/maier2018/):
  - drugs.xlsx        (original MOESM3, sheet "S1a. Prestwick_Libery")
  - hits_matrix.xlsx   (original MOESM5, sheet "S3a. Adjusted p-values")

A "hit" is a drug-strain pair with FDR-adjusted p < 0.01. Only hits
are stored: effect_direction is always 'decreases' (growth inhibition).
The real adjusted p-value is kept in a new column rather than discarded.

Usage:
    pip install openpyxl
    python scripts/16_load_drug_interventions.py
"""

import re
import sqlite3
import time
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import load_workbook
import requests

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_PATH = PROJECT_ROOT / "db" / "microbiome.db"
DATA_DIR = PROJECT_ROOT / "data" / "raw" / "maier2018"

DRUGS_FILE = DATA_DIR / "drugs.xlsx"
MATRIX_FILE = DATA_DIR / "hits_matrix.xlsx"

STUDY_ID = "DOI:10.1038/nature25979"
FDR_THRESHOLD = 0.01
EUTILS_BASE = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
REQUEST_DELAY = 0.4


def search_taxid(name: str):
    resp = requests.get(
        f"{EUTILS_BASE}/esearch.fcgi",
        params={"db": "taxonomy", "term": name, "retmode": "json"},
        timeout=15,
    )
    resp.raise_for_status()
    id_list = resp.json().get("esearchresult", {}).get("idlist", [])
    return id_list[0] if id_list else None


def fetch_taxon_record(taxid: str) -> dict:
    resp = requests.get(
        f"{EUTILS_BASE}/efetch.fcgi",
        params={"db": "taxonomy", "id": taxid, "rettype": "xml"},
        timeout=15,
    )
    resp.raise_for_status()
    root = ET.fromstring(resp.content)
    taxon = root.find("Taxon")
    if taxon is None:
        raise ValueError(f"no <Taxon> element for taxid {taxid}")
    record = {
        "taxid": taxon.findtext("TaxId"),
        "name": taxon.findtext("ScientificName"),
        "rank": (taxon.findtext("Rank") or "no rank").lower(),
        "parent_taxid": taxon.findtext("ParentTaxId"),
        "synonyms": [],
    }
    other_names = taxon.find("OtherNames")
    if other_names is not None:
        for tag in ("GenbankSynonym", "Synonym", "EquivalentName"):
            for el in other_names.findall(tag):
                if el.text:
                    record["synonyms"].append(el.text)
    return record


def lookup_and_insert_organism(conn: sqlite3.Connection, species_name: str):
    taxid = search_taxid(species_name)
    time.sleep(REQUEST_DELAY)
    if not taxid:
        return None
    record = fetch_taxon_record(taxid)
    time.sleep(REQUEST_DELAY)
    conn.execute(
        """
        INSERT OR IGNORE INTO organism
            (ncbi_taxid, name, rank, parent_taxid, source_db, last_verified)
        VALUES (?, ?, ?, ?, ?, date('now'))
        """,
        (record["taxid"], record["name"], record["rank"], record["parent_taxid"], "NCBI Taxonomy"),
    )
    conn.execute("DELETE FROM organism_synonym WHERE ncbi_taxid = ?", (record["taxid"],))
    for syn in record["synonyms"]:
        conn.execute(
            "INSERT INTO organism_synonym (ncbi_taxid, synonym_name) VALUES (?, ?)",
            (record["taxid"], syn),
        )
    conn.commit()
    print(f"  [live lookup] {species_name} -> {record['name']}, taxid {record['taxid']}")
    return record["taxid"]


NT_CODE_SPECIES = {
    "NT5001": "Phocaeicola vulgatus",          # was: Bacteroides vulgatus
    "NT5002": "Bacteroides uniformis",
    "NT5003": "Bacteroides fragilis",
    "NT5004": "Bacteroides thetaiotaomicron",
    "NT5006": "Clostridium ramosum",
    "NT5009": "Agathobacter rectalis",         # was: Eubacterium rectale
    "NT5011": "Roseburia intestinalis",
    "NT5017": "Veillonella parvula",
    "NT5019": "Segatella copri",               # was: Prevotella copri
    "NT5021": "Akkermansia muciniphila",
    "NT5022": "Bifidobacterium adolescentis",
    "NT5024": "Eggerthella lenta",
    "NT5025": "Fusobacterium nucleatum",
    "NT5026": "Enterocloster bolteae",         # was: Clostridium bolteae
    "NT5028": "Bifidobacterium longum",
    "NT5032": "Clostridium perfringens",
    "NT5033": "Bacteroides fragilis",
    "NT5036": "Bilophila wadsworthia",
    "NT5037": "Lacrimispora saccharolytica",   # was: Clostridium saccharolyticum
    "NT5038": "Streptococcus salivarius",
    "NT5042": "Lacticaseibacillus paracasei",  # was: Lactobacillus paracasei
    "NT5045": "Ruminococcus bromii",
    "NT5046": "Ruminococcus gnavus",
    "NT5047": "Mediterraneibacter torques",    # was: Ruminococcus torques
    "NT5048": "Allocoprococcus comes",         # was: Coprococcus comes
    "NT5050": "Bacteroides caccae",
    "NT5054": "Bacteroides ovatus",
    "NT5064": "Bacteroides xylanisolvens",
    "NT5069": "Blautia obeum",
    "NT5071": "Parabacteroides merdae",
    "NT5072": "Streptococcus parasanguinis",
    "NT5073": "Collinsella aerofaciens",
    "NT5074": "Parabacteroides distasonis",
    "NT5075": "Lachnospira eligens",           # was: Eubacterium eligens
    "NT5076": "Dorea formicigenerans",
    "NT5077": "Escherichia coli",
    "NT5078": "Escherichia coli",
    "NT5079": "Roseburia hominis",
    "NT5081": "Odoribacter splanchnicus",
    "NT5083": "Clostridium difficile",
}


def ensure_columns(conn: sqlite3.Connection) -> None:
    cols = {row[1] for row in conn.execute("PRAGMA table_info(microbe_intervention)")}
    if "adjusted_pvalue" not in cols:
        conn.execute("ALTER TABLE microbe_intervention ADD COLUMN adjusted_pvalue REAL")
    conn.commit()


def build_organism_lookup(conn) -> list:
    pairs = []
    for taxid, name in conn.execute("SELECT ncbi_taxid, name FROM organism"):
        pairs.append((name.lower(), taxid))
    for taxid, syn in conn.execute("SELECT ncbi_taxid, synonym_name FROM organism_synonym"):
        pairs.append((syn.lower(), taxid))
    return pairs


def find_taxid(pairs: list, species_name: str):
    target = species_name.lower()
    for stored, taxid in pairs:
        if stored == target:
            return taxid
    for stored, taxid in pairs:
        if stored.startswith(target + " "):
            return taxid
    return None


def load_drugs(conn: sqlite3.Connection) -> int:
    wb = load_workbook(DRUGS_FILE, read_only=True)
    ws = wb["S1a. Prestwick_Libery"]
    rows = ws.iter_rows(values_only=True)
    next(rows)

    count = 0
    for row in rows:
        prestwick_id, name = row[0], row[1]
        if not prestwick_id or not name:
            continue
        conn.execute(
            "INSERT OR IGNORE INTO intervention (intervention_id, type, name, source) VALUES (?, ?, ?, ?)",
            (prestwick_id, "drug", name, "Maier et al. 2018 (Prestwick Chemical Library)"),
        )
        count += 1
    conn.commit()
    return count


def main():
    if not DRUGS_FILE.exists() or not MATRIX_FILE.exists():
        raise SystemExit(f"Expected {DRUGS_FILE} and {MATRIX_FILE} -- copy them there first.")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = OFF")
    ensure_columns(conn)

    conn.execute(
        "INSERT OR IGNORE INTO study (study_id, study_type, detection_method) VALUES (?, ?, ?)",
        (STUDY_ID, "in_vitro", "high-throughput growth inhibition screen (AUC, FDR<0.01)"),
    )

    n_drugs = load_drugs(conn)
    print(f"Loaded {n_drugs} drugs into intervention table.")

    organism_lookup = build_organism_lookup(conn)
    nt_to_taxid = {}
    unmatched_species = set()
    for nt_code, species in NT_CODE_SPECIES.items():
        taxid = find_taxid(organism_lookup, species)
        if taxid:
            nt_to_taxid[nt_code] = taxid
        else:
            unmatched_species.add((nt_code, species))

    if unmatched_species:
        print(f"{len(unmatched_species)} species not found locally -- trying live NCBI lookup:")
        for nt_code, species in unmatched_species:
            taxid = lookup_and_insert_organism(conn, species)
            if taxid:
                nt_to_taxid[nt_code] = taxid
            else:
                print(f"  [skip] no taxid found for '{species}'")

    print(f"\nMatched {len(nt_to_taxid)} of {len(NT_CODE_SPECIES)} screened strains to our organism table.")

    wb = load_workbook(MATRIX_FILE, read_only=True)
    ws = wb["S3a. Adjusted p-values"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)

    col_taxid = {}
    for i, col_name in enumerate(header[4:], start=4):
        match = re.search(r"\((NT\d+)\)", col_name or "")
        if match and match.group(1) in nt_to_taxid:
            col_taxid[i] = nt_to_taxid[match.group(1)]

    inserted, skipped_existing = 0, 0
    for row in rows:
        prestwick_id = row[0]
        if not prestwick_id:
            continue
        for i, taxid in col_taxid.items():
            try:
                pvalue = float(row[i]) if row[i] is not None else None
            except (TypeError, ValueError):
                pvalue = None
            if pvalue is None or pvalue >= FDR_THRESHOLD:
                continue

            existing = conn.execute(
                "SELECT 1 FROM microbe_intervention WHERE ncbi_taxid = ? AND intervention_id = ?",
                (taxid, prestwick_id),
            ).fetchone()
            if existing:
                skipped_existing += 1
                continue

            conn.execute(
                """
                INSERT INTO microbe_intervention
                    (ncbi_taxid, intervention_id, effect_direction, evidence_level,
                     study_id, adjusted_pvalue)
                VALUES (?, ?, 'decreases', 'in vitro growth inhibition screen', ?, ?)
                """,
                (taxid, prestwick_id, STUDY_ID, pvalue),
            )
            inserted += 1

    conn.commit()
    conn.execute("PRAGMA foreign_keys = ON")
    issues = conn.execute("PRAGMA foreign_key_check").fetchall()
    conn.close()

    print(
        f"\nDone. {inserted} drug-organism inhibition hits inserted, "
        f"{skipped_existing} duplicate (organism, drug) pairs skipped "
        f"(e.g. two E. coli strains both hit by the same drug)."
    )
    if issues:
        print(f"WARNING: {len(issues)} foreign key issues:")
        for i in issues[:10]:
            print(f"  {i}")
    else:
        print("Foreign key integrity check passed.")


if __name__ == "__main__":
    main()
