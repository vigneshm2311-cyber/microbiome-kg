import sqlite3
from pathlib import Path
from openpyxl import load_workbook

DB_PATH = Path("db/microbiome.db")
MATRIX_FILE = Path("data/raw/maier2018/hits_matrix.xlsx")
TARGET_TAXID = "84030"
FDR_THRESHOLD = 0.01

conn = sqlite3.connect(DB_PATH)

wb = load_workbook(MATRIX_FILE, read_only=True)
ws = wb["S3a. Adjusted p-values"]
rows = ws.iter_rows(values_only=True)
header = next(rows)
idx = next(i for i, col in enumerate(header) if col and "NT5037" in str(col))
print(f"Reading column {idx} for NT5037 (taxid {TARGET_TAXID})")

hits_in_file, would_insert, already_existed, errors = 0, 0, 0, 0

for row_num, row in enumerate(rows, start=2):
    prestwick_id = row[0]
    if not prestwick_id:
        continue
    raw_value = row[idx]
    try:
        pvalue = float(raw_value) if raw_value is not None else None
    except (TypeError, ValueError):
        pvalue = None

    if pvalue is None or pvalue >= FDR_THRESHOLD:
        continue

    hits_in_file += 1
    existing = conn.execute(
        "SELECT 1 FROM microbe_intervention WHERE ncbi_taxid = ? AND intervention_id = ?",
        (TARGET_TAXID, prestwick_id),
    ).fetchone()
    if existing:
        already_existed += 1
        if already_existed <= 3:
            print(f"  row {row_num}: {prestwick_id} p={pvalue:.2e} -- ALREADY IN DB")
    else:
        would_insert += 1
        if would_insert <= 3:
            print(f"  row {row_num}: {prestwick_id} p={pvalue:.2e} -- NOT in DB, would insert")

print(f"\nTotal real hits in file for this column: {hits_in_file}")
print(f"Already in microbe_intervention: {already_existed}")
print(f"Would insert (genuinely new): {would_insert}")

sample_id = None
for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
    sample_id = row[0]
print(f"\nintervention table has this drug? {conn.execute('SELECT * FROM intervention WHERE intervention_id=?', (sample_id,)).fetchone()}")

conn.close()
